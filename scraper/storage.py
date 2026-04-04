import json
import hashlib
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .logger import general_log

# Setup folders safely
OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

CACHE_DIR = Path(".cache")
CACHE_DIR.mkdir(exist_ok=True)

# Professional column order for product tables
PRODUCT_COLUMNS = [
    "#", "Product Name", "Brand", "Sale Price", "Original Price",
    "Discount (%)", "Rating", "Product URL", "Image URL",
]


def cache_html(html: str, url: str) -> None:
    """Saves raw HTML for debugging, using MD5 hash of URL as filename."""
    if not html:
        return
    url_hash = hashlib.md5(url.encode('utf-8')).hexdigest()
    filename = CACHE_DIR / f"{url_hash}.html"
    try:
        with open(filename, "w", encoding="utf-8") as f:
            f.write(html)
        general_log.info(f"Raw HTML cached to {filename}")
    except Exception as e:
        general_log.error(f"Failed to cache HTML: {e}")


def _get_dataframe(cleaned_data: dict) -> pd.DataFrame:
    """
    Builds a Pandas DataFrame with professional column names.
    Enforces column order from PRODUCT_COLUMNS.
    """
    products = cleaned_data.get("products", [])
    summary = cleaned_data.get("summary", {})

    if products:
        df = pd.DataFrame(products)
        # Enforce column order, keeping only existing columns
        ordered = [c for c in PRODUCT_COLUMNS if c in df.columns]
        extra = [c for c in df.columns if c not in PRODUCT_COLUMNS]
        df = df[ordered + extra]
        return df
    else:
        row = {k: v for k, v in summary.items() if k != "Error Message" or v}
        return pd.DataFrame([row])


# ─── CSV ──────────────────────────────────────────────────────────────────────

def save_to_csv(cleaned_data: dict, filename: str = "scraped_data.csv") -> None:
    """Saves data to a professional CSV with clear column headers."""
    filepath = OUTPUT_DIR / filename
    try:
        df = _get_dataframe(cleaned_data)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        general_log.info(f"CSV saved: {filepath} ({len(df)} rows × {len(df.columns)} columns)")
    except Exception as e:
        general_log.error(f"Failed to save CSV: {e}")


# ─── JSON ─────────────────────────────────────────────────────────────────────

def save_to_json(data, filename: str = "scraped_data.json") -> None:
    """
    Saves data to clean, readable JSON.
    Handles lists of profiles, single profiles, or listings.
    """
    filepath = OUTPUT_DIR / filename
    try:
        # Check if it's a list or Business Profile
        if isinstance(data, list):
            output = data
        elif isinstance(data, dict) and "business" in data:
            output = data
        else:
            products = data.get("products", [])
            summary = data.get("summary", {})
            output = {
                "scrape_summary": {
                    "page_title": summary.get("Page Title", ""),
                    "url": summary.get("URL", ""),
                    "description": summary.get("Description", ""),
                    "total_items_extracted": len(products),
                },
                "products": products,
            }

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=4, ensure_ascii=False)
        general_log.info(f"JSON saved: {filepath}")
    except Exception as e:
        general_log.error(f"Failed to save JSON: {e}")


# ─── TXT ──────────────────────────────────────────────────────────────────────

def save_to_txt(cleaned_data: dict, filename: str = "scraped_data.txt") -> None:
    """Saves data in a neatly formatted human-readable text report."""
    filepath = OUTPUT_DIR / filename
    try:
        summary = cleaned_data.get("summary", {})
        products = cleaned_data.get("products", [])

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("=" * 80 + "\n")
            f.write("  SCRAPING RESULTS — STRUCTURED DATA REPORT\n")
            f.write("=" * 80 + "\n\n")

            f.write(f"  Page:        {summary.get('Page Title', 'N/A')}\n")
            f.write(f"  URL:         {summary.get('URL', 'N/A')}\n")
            f.write(f"  Description: {summary.get('Description', 'N/A')[:120]}\n")
            f.write(f"  JSON-LD:     {summary.get('Has JSON-LD', 'No')}\n\n")

            if products:
                f.write(f"  Total Products Extracted: {len(products)}\n")
                f.write("-" * 80 + "\n\n")

                for product in products:
                    idx = product.get("#", "?")
                    f.write(f"  [{idx}] {product.get('Product Name', 'N/A')}\n")
                    f.write(f"       Brand:          {product.get('Brand', '—')}\n")
                    f.write(f"       Sale Price:     {product.get('Sale Price', '—')}\n")
                    f.write(f"       Original Price: {product.get('Original Price', '—')}\n")
                    f.write(f"       Discount:       {product.get('Discount (%)', '—')}\n")
                    f.write(f"       Rating:         {product.get('Rating', '—')}\n")
                    f.write(f"       Product URL:    {product.get('Product URL', '—')}\n")
                    f.write(f"       Image URL:      {product.get('Image URL', '—')}\n")
                    f.write("\n")
            else:
                f.write("  No product cards detected on this page.\n")

            f.write("=" * 80 + "\n")

        general_log.info(f"TXT saved: {filepath}")
    except Exception as e:
        general_log.error(f"Failed to save TXT: {e}")


# ─── Excel ────────────────────────────────────────────────────────────────────

def save_to_excel(cleaned_data: dict, filename: str = "scraped_data.xlsx") -> None:
    """Saves data to a professionally formatted Excel file with styled headers."""
    filepath = OUTPUT_DIR / filename
    try:
        df = _get_dataframe(cleaned_data)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Scraped Data", startrow=1)

            workbook = writer.book
            worksheet = writer.sheets["Scraped Data"]

            # ── Style: Header row ──
            header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Write styled headers (row 2, since row 1 is for title)
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=2, column=col_idx)
                cell.value = col_name
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # ── Style: Title row ──
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.value = f"Scraped Data — {cleaned_data.get('summary', {}).get('Page Title', 'Report')}"
            title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F3864")
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

            # ── Style: Data rows ──
            data_font = Font(name="Calibri", size=10)
            data_align = Alignment(vertical="center", wrap_text=False)
            alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

            for row_idx in range(3, len(df) + 3):
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border
                    # Alternating row colors
                    if row_idx % 2 == 1:
                        cell.fill = alt_fill

            # ── Auto-adjust column widths ──
            for col_idx, col_name in enumerate(df.columns, 1):
                max_len = max(
                    df[col_name].astype(str).map(len).max() if len(df) > 0 else 0,
                    len(col_name)
                ) + 3
                # Cap between 8 and 50 chars
                col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                worksheet.column_dimensions[col_letter].width = max(8, min(max_len, 50))

            # Freeze the header rows
            worksheet.freeze_panes = "A3"

        general_log.info(f"Excel saved: {filepath} ({len(df)} rows × {len(df.columns)} columns)")
    except Exception as e:
        general_log.error(f"Failed to save Excel: {e}")
